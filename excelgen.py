from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_timetable(schedule: dict, title: str = "Toetsweek", output_path: str = "timetable.xlsx", wb=None, sheet_name="Timetable"):
    """
    schedule: dict of shape {day_name: {period_num: {"exams": [Exam, ...], "rooms": [...]}}}
    title: spreadsheet title shown in the header
    output_path: where to save the .xlsx
    """

    # --- collect all rooms, sorted ---
    all_rooms = sorted({
        exam.room
        for day_data in schedule.values()
        for period_data in day_data.values()
        for exam in period_data["exams"]
    })

    if wb is None:
        wb = Workbook()
        ws = ws
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    # ── Styles ──────────────────────────────────────────────────────────────
    BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")   # room header
    GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")   # day separator row
    LIGHT_FILL  = PatternFill("solid", fgColor="F2F2F2")   # sub-header row
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

    thin  = Side(style="thin",   color="000000")
    med   = Side(style="medium", color="000000")

    def thin_border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def med_border():
        return Border(left=med, right=med, top=med, bottom=med)

    def bold_font(size=9, color="000000"):
        return Font(name="Arial", bold=True, size=size, color=color)

    def reg_font(size=9, color="000000"):
        return Font(name="Arial", size=size, color=color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    # ── Layout constants ─────────────────────────────────────────────────────
    FIXED_COLS = 3          # col A=day, B=date, C=period
    ROOM_COLS  = 5          # start/einde, uur, klas, vak, surv.
    NUM_ROOMS  = len(all_rooms)
    EXTRA_COLS = 2          # gang, reserve

    total_cols = FIXED_COLS + NUM_ROOMS * ROOM_COLS + EXTRA_COLS

    def room_col_start(room_idx):
        return FIXED_COLS + 1 + room_idx * ROOM_COLS   # 1-based

    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 8   # day
    ws.column_dimensions[get_column_letter(2)].width = 10  # date
    ws.column_dimensions[get_column_letter(3)].width = 4   # period
    for r in range(NUM_ROOMS):
        base = room_col_start(r)
        ws.column_dimensions[get_column_letter(base)].width     = 12  # start/einde
        ws.column_dimensions[get_column_letter(base+1)].width   = 4   # uur
        ws.column_dimensions[get_column_letter(base+2)].width   = 7   # klas
        ws.column_dimensions[get_column_letter(base+3)].width   = 7   # vak
        ws.column_dimensions[get_column_letter(base+4)].width   = 9   # surv.
    gang_col    = FIXED_COLS + NUM_ROOMS * ROOM_COLS + 1
    reserve_col = gang_col + 1
    ws.column_dimensions[get_column_letter(gang_col)].width    = 6
    ws.column_dimensions[get_column_letter(reserve_col)].width = 8

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=FIXED_COLS + NUM_ROOMS * ROOM_COLS)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Arial", bold=True, size=14)
    c.alignment = center()

    ws.cell(row=1, column=gang_col,    value="gang").font    = bold_font()
    ws.cell(row=1, column=reserve_col, value="reserve").font = bold_font()
    ws.cell(row=1, column=gang_col).alignment    = center()
    ws.cell(row=1, column=reserve_col).alignment = center()

    current_row = 2

    DAY_LABELS = {
        "Mon": "maandag", "Tu": "dinsdag", "Wed": "woensdag",
        "Thu": "donderdag", "Fri": "vrijdag", "Sat": "zaterdag", "Sun": "zondag"
    }

    for day_key, day_data in schedule.items():
        # FIX 2: unwrap {"exams": [...]} when checking for used periods
        used_periods = sorted(p for p, period_data in day_data.items() if period_data["exams"])
        if not used_periods:
            continue
        max_period = max(used_periods)

        # ── Room header row ──────────────────────────────────────────────────
        header_row = current_row
        for r_idx, room in enumerate(all_rooms):
            base = room_col_start(r_idx)
            ws.merge_cells(start_row=header_row, start_column=base,
                           end_row=header_row, end_column=base + ROOM_COLS - 1)
            c = ws.cell(row=header_row, column=base, value=f"lokaal {room}")
            c.font      = bold_font(color="1F4E79")
            c.fill      = BLUE_FILL
            c.alignment = center()
            c.border    = thin_border()

        for col in range(1, FIXED_COLS + 1):
            ws.cell(row=header_row, column=col).fill   = WHITE_FILL
            ws.cell(row=header_row, column=col).border = thin_border()

        current_row += 1

        # ── Sub-header row ───────────────────────────────────────────────────
        sub_row = current_row
        sub_headers = ["duur", "uur", "klas", "vak", "surv."]
        for r_idx in range(NUM_ROOMS):
            base = room_col_start(r_idx)
            for i, lbl in enumerate(sub_headers):
                c = ws.cell(row=sub_row, column=base + i, value=lbl)
                c.font      = bold_font(size=8)
                c.fill      = LIGHT_FILL
                c.alignment = center()
                c.border    = thin_border()

        for col in range(1, FIXED_COLS + 1):
            ws.cell(row=sub_row, column=col).fill   = LIGHT_FILL
            ws.cell(row=sub_row, column=col).border = thin_border()

        current_row += 1

        # ── Data rows: one per period ────────────────────────────────────────
        period_rows = []
        for period in range(1, max_period + 1):
            row = current_row
            period_rows.append(row)
            ws.row_dimensions[row].height = 18

            # FIX 3: unwrap {"exams": [...]} when getting exams for a period
            exams_this_period = day_data.get(period, {}).get("exams", [])
            exam_by_room = {e.room: e for e in exams_this_period}

            c = ws.cell(row=row, column=3, value=period)
            c.font      = reg_font()
            c.alignment = center()
            c.border    = thin_border()

            for col in [1, 2]:
                ws.cell(row=row, column=col).border = thin_border()

            for r_idx, room in enumerate(all_rooms):
                base = room_col_start(r_idx)
                exam = exam_by_room.get(room)

                start_einde = exam.duration if exam else ""
                uur         = str(period)   if exam else ""
                klas        = exam.group_id.split("_")[0] if exam else ""
                vak         = exam.subject  if exam else ""
                surv        = exam.supervisor if exam else ""

                vals = [start_einde, uur, klas, vak, surv]
                for i, val in enumerate(vals):
                    c = ws.cell(row=row, column=base + i, value=val)
                    c.font      = reg_font()
                    c.alignment = center()
                    c.border    = thin_border()

            current_row += 1

        # ── Merge day label (col A) ──────────────────────────────────────────
        day_start_row = header_row
        day_end_row   = current_row - 1

        if day_end_row > day_start_row:
            ws.merge_cells(start_row=day_start_row, start_column=1,
                           end_row=day_end_row, end_column=1)
        c = ws.cell(row=day_start_row, column=1,
                    value=DAY_LABELS.get(day_key, day_key))
        c.font      = bold_font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                text_rotation=90, wrap_text=False)
        c.border    = med_border()

        # ── Merge date (col B) ───────────────────────────────────────────────
        if period_rows:
            ws.merge_cells(start_row=period_rows[0], start_column=2,
                           end_row=period_rows[-1], end_column=2)
        c = ws.cell(row=period_rows[0] if period_rows else header_row, column=2)
        c.font      = reg_font(size=8)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                text_rotation=90)
        c.border    = thin_border()

        # ── Day separator row ────────────────────────────────────────────────
        sep_row = current_row
        for col in range(1, total_cols + 1):
            c = ws.cell(row=sep_row, column=col)
            c.fill   = GRAY_FILL
            c.border = Border(top=med, bottom=med)
        ws.row_dimensions[sep_row].height = 6
        current_row += 1

    if output_path:
        wb.save(output_path)
        print(f"Saved: {output_path}")
    return wb
